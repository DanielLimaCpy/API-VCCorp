import pandas as pd
import json
import openpyxl

def criar_planilha_excel(json_file, nome_arquivo_excel):
    with open(json_file, 'r') as file:
        dados = json.load(file)

    def remove_colchetes(value):
        if isinstance(value, list):
            return ', '.join(map(str, value))
        return value


    with pd.ExcelWriter(nome_arquivo_excel, engine='openpyxl') as writer:
    
        secoes = ["alunos", "grupos", "turmas", "ciclos", "notas"]
        for secao in secoes:
           
            secao_data = dados.get(secao, {})

         
            if secao == "turmas":
                turmas_dict = {}
                for turma_id, turma_data in secao_data.items():
                    ciclo_ids = [ciclo["id"] for ciclo in turma_data.get("ciclos", [])]
                    turma_data["ciclos"] = ', '.join(ciclo_ids)
                    turmas_dict[turma_id] = turma_data

                df_secao = pd.DataFrame.from_dict(turmas_dict, orient='index').applymap(remove_colchetes)
            else:
                df_secao = pd.DataFrame.from_dict(secao_data, orient='index').applymap(remove_colchetes)

            # Salvando na planilha Excel
            df_secao.to_excel(writer, index=False, sheet_name=secao.capitalize())

            sheet = writer.sheets[secao.capitalize()]

            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                series = pd.Series(column)

                if series.dtype == 'O':  # Verifica se a coluna é de tipo string
                    max_length = max(
                        series.astype(str).apply(len).max(),  # Este é o comprimento do cabeçalho da coluna
                        series.str.len().max()  # Isso é para o comprimento do conteúdo máximo da coluna
                    )

                # Definindo limites mínimo e máximo para a largura da coluna
                min_width = 8  # Largura mínima
                max_width = 60  # Largura máxima

                # Ajustando a largura da coluna para o máximo encontrado dentro dos limites
                adjusted_width = max(min(max_length + 2, max_width), min_width)
                sheet.column_dimensions[chr(65 + column[0].column)].width = adjusted_width

            # Centralizando todas as células
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    print(f'Arquivo Excel "{nome_arquivo_excel}" criado com sucesso.')





